from __future__ import annotations

import sqlite3
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from asset_hub.catalog.db import AssetRow, Catalog
from asset_hub.catalog.ignore import should_ignore
from asset_hub.catalog.index import index_library
from asset_hub.config import Settings, ensure_data_dirs, get_settings
from asset_hub.jobs import JobStore
from asset_hub.pack.excel import (
    ExcelRow,
    deduplicate_rows,
    match_assets_for_rows,
    read_excel_rows,
)
from asset_hub.pack.ziputil import zip_paths
from asset_hub.pack.rules import PackRuleStore
from asset_hub.sync.provider import ExternalManifestItem, ManifestItem
from asset_hub.sync.main import sync_once


@pytest.fixture()
def settings(tmp_path, monkeypatch):
    data = tmp_path / "data"
    lib = tmp_path / "library"
    lib.mkdir()
    cfg = tmp_path / "config.yaml"
    cfg.write_text(
        f"""
library_root: {lib}
data_root: {data}
local_only: true
provider: mock
api:
  host: 127.0.0.1
  port: 18080
sync:
  kinds: [finalized]
  ignore_globs: ["Thumbs.db", "desktop.ini", "._*"]
""".strip(),
        encoding="utf-8",
    )
    monkeypatch.setenv("ASSET_HUB_CONFIG", str(cfg))
    monkeypatch.delenv("ASSET_HUB_DATA", raising=False)
    monkeypatch.delenv("ASSET_HUB_LIBRARY", raising=False)
    monkeypatch.setenv("ASSET_HUB_X_ACCEL", "0")
    get_settings.cache_clear()
    s = ensure_data_dirs(get_settings())
    yield s
    get_settings.cache_clear()


def test_ignore_rules():
    assert should_ignore("Thumbs.db")
    assert should_ignore("desktop.ini")
    assert should_ignore("._hidden")
    assert not should_ignore("photo.jpg")


def test_catalog_search_sku(settings):
    cat = Catalog(settings)
    cat.upsert_asset(
        AssetRow(
            asset_id="a1",
            kind="finalized",
            file_name="HQT10001-x.jpg",
            sku_code="HQT10001",
            sku_name="测试品",
            local_path="/tmp/x",
            status="ready",
        )
    )
    hits = cat.search("HQT10001", kind="finalized")
    assert len(hits[0]) == 1
    assert hits[0][0].asset_id == "a1"


def test_catalog_search_normalizes_and_reads_legacy_trailing_punctuation(settings):
    cat = Catalog(settings)
    asset = AssetRow(
        asset_id="external:hyphenated-sku",
        kind="external",
        file_name="HSC11108-——常规KT板.jpg",
        local_path="/tmp/HSC11108.jpg",
        status="ready",
    )
    cat.upsert_asset(asset)
    with cat.connect() as conn:
        tokens = {
            row["token"]
            for row in conn.execute(
                "SELECT token FROM sku_tokens WHERE asset_id=?", (asset.asset_id,)
            ).fetchall()
        }
        assert "HSC11108" in tokens

        # Simulate a row indexed before trailing punctuation normalization.
        conn.execute("DELETE FROM sku_tokens WHERE asset_id=?", (asset.asset_id,))
        conn.execute(
            "INSERT INTO sku_tokens(token,asset_id) VALUES (?,?)",
            ("HSC11108-", asset.asset_id),
        )

    hits, total = cat.search("HSC11108", kind="external")
    assert total == 1
    assert [row.asset_id for row in hits] == [asset.asset_id]


def test_catalog_search_hides_library_under_external_overlay(settings):
    cat = Catalog(settings)
    path = "1/KT/HSC38201-套装/上图.jpg"
    cat.upsert_assets(
        [
            AssetRow(
                asset_id="lib:overlay",
                kind="library",
                file_name="上图.jpg",
                local_path="/tmp/library.jpg",
                virtual_path=path,
                status="ready",
            ),
            AssetRow(
                asset_id="external:overlay",
                kind="external",
                file_name="上图.jpg",
                local_path="/tmp/external.jpg",
                virtual_path=path,
                status="ready",
            ),
        ]
    )

    hits, total = cat.search("HSC38201")
    assert total == 1
    assert [row.asset_id for row in hits] == ["external:overlay"]

def test_catalog_bulk_upsert(settings):
    cat = Catalog(settings)
    rows = [
        AssetRow(
            asset_id=f"bulk:{index}",
            kind="library",
            file_name=f"HQT-BULK-{index}.jpg",
            sku_code=f"HQT-BULK-{index}",
            status="ready",
        )
        for index in range(3)
    ]

    assert cat.upsert_assets(rows) == 3
    assert cat.count_ready("library") == 3


def test_unified_search_ready_only_and_current_first(settings, tmp_path):
    cat = Catalog(settings)
    library_file = tmp_path / "library-HQT20001.jpg"
    current_file = tmp_path / "current-HQT20001.jpg"
    library_file.write_bytes(b"library")
    current_file.write_bytes(b"current")
    cat.upsert_asset(
        AssetRow(
            asset_id="lib:HQT20001.jpg",
            kind="library",
            file_name=library_file.name,
            sku_code="HQT20001",
            local_path=str(library_file),
            status="ready",
        )
    )
    cat.upsert_asset(
        AssetRow(
            asset_id="finalized:20001",
            task_asset_id=20001,
            kind="finalized",
            file_name=current_file.name,
            sku_code="HQT20001",
            local_path=str(current_file),
            status="ready",
        )
    )
    cat.upsert_asset(
        AssetRow(
            asset_id="finalized:pending",
            task_asset_id=20002,
            kind="finalized",
            file_name="pending-HQT20001.jpg",
            sku_code="HQT20001",
            status="pending",
        )
    )

    hits, total = cat.search("HQT20001")
    assert total == 2
    assert [hit.asset_id for hit in hits] == [
        "finalized:20001",
        "lib:HQT20001.jpg",
    ]


def test_unified_match_prefers_current_and_falls_back(settings, tmp_path):
    cat = Catalog(settings)
    current = tmp_path / "current.jpg"
    fallback_same = tmp_path / "fallback-same.jpg"
    fallback_only = tmp_path / "fallback-only.jpg"
    current.write_bytes(b"current")
    fallback_same.write_bytes(b"fallback")
    fallback_only.write_bytes(b"fallback-only")
    for asset in (
        AssetRow(
            asset_id="finalized:30001",
            task_asset_id=30001,
                kind="finalized",
                file_name=current.name,
                file_size=current.stat().st_size,
                sku_code="HQT30001",
            local_path=str(current),
            status="ready",
        ),
        AssetRow(
            asset_id="lib:HQT30001.jpg",
                kind="library",
                file_name=fallback_same.name,
                file_size=fallback_same.stat().st_size,
                sku_code="HQT30001",
            local_path=str(fallback_same),
            status="ready",
        ),
        AssetRow(
            asset_id="lib:HQT30002.jpg",
                kind="library",
                file_name=fallback_only.name,
                file_size=fallback_only.stat().st_size,
                sku_code="HQT30002",
            local_path=str(fallback_only),
            status="ready",
        ),
    ):
        cat.upsert_asset(asset)

    matched, missing = match_assets_for_rows(
        cat,
        [
            ExcelRow(row_index=2, sku_code="HQT30001"),
            ExcelRow(row_index=3, sku_code="HQT30002"),
        ],
    )
    assert not missing
    assert matched[0]["selection_policy"] == "preferred_current"
    assert [asset.asset_id for asset in matched[0]["assets"]] == ["finalized:30001"]
    assert matched[1]["selection_policy"] == "library_fallback"
    assert [asset.asset_id for asset in matched[1]["assets"]] == ["lib:HQT30002.jpg"]


def test_sync_mock_and_ready(settings):
    stats = sync_once()
    assert stats["written"] > 0 or stats["fetched"] > 0
    cat = Catalog(settings)
    assert cat.count_ready("finalized") > 0
    assert cat.is_finalized_ready()


def test_index_library(settings):
    root = settings.library_root
    (root / "skuA").mkdir()
    (root / "skuA" / "a.jpg").write_bytes(b"abc")
    (root / "Thumbs.db").write_bytes(b"x")
    n = index_library(Catalog(settings), root, settings.sync.ignore_globs)
    assert n == 1
    hits = Catalog(settings).search("a.jpg", kind="library")
    assert len(hits[0]) == 1


def test_index_skips_unchanged_and_tombstones_removed(settings):
    root = settings.library_root
    path = root / "SKU90001.jpg"
    path.write_bytes(b"image")
    catalog = Catalog(settings)
    assert index_library(catalog, root, settings.sync.ignore_globs) == 1

    writes: list[int] = []
    original = catalog.upsert_assets

    def counted(assets):
        writes.append(len(assets))
        return original(assets)

    catalog.upsert_assets = counted  # type: ignore[method-assign]
    assert index_library(catalog, root, settings.sync.ignore_globs) == 1
    assert writes == []

    path.unlink()
    assert index_library(catalog, root, settings.sync.ignore_globs) == 0
    asset = catalog.get_asset("lib:SKU90001.jpg")
    assert asset is not None and asset.deleted == 1 and asset.status == "tombstone"


def test_index_searches_sku_from_product_directory_with_generic_names(settings):
    root = settings.library_root
    product = root / "水晶标" / "HSC36004——蔡谦-常规水晶标-喜字酒杯款-直径45cm"
    other = root / "水晶标" / "HSC33778——叶真-常规水晶标-大红喜字款-直径45cm"
    product.mkdir(parents=True)
    other.mkdir(parents=True)
    (product / "第一张【25x35cm】.jpg").write_bytes(b"one")
    (product / "第二张【25x35cm】.jpg").write_bytes(b"two")
    (other / "第一张【25x35cm】.jpg").write_bytes(b"other")

    catalog = Catalog(settings)
    assert index_library(catalog, root, settings.sync.ignore_globs) == 3
    hits, total = catalog.search("HSC36004", limit=20)
    assert total == 2
    assert {hit.file_name for hit in hits} == {
        "第一张【25x35cm】.jpg",
        "第二张【25x35cm】.jpg",
    }


def test_search_does_not_hide_current_assets_with_same_generic_filename(settings):
    catalog = Catalog(settings)
    first = settings.finalized_dir / "1" / "第一块【42x50cm】.jpg"
    second = settings.finalized_dir / "2" / "第一块【42x50cm】.jpg"
    first.parent.mkdir(parents=True)
    second.parent.mkdir(parents=True)
    first.write_bytes(b"first-content")
    second.write_bytes(b"second-content")
    catalog.upsert_assets(
        [
            AssetRow(
                asset_id="finalized:1",
                task_asset_id=1,
                kind="finalized",
                file_name=first.name,
                file_size=first.stat().st_size,
                sku_code="CGK000001",
                sku_name="商品一/3个装",
                local_path=str(first),
                status="ready",
                virtual_path=f"CGK000001/{first.name}",
            ),
            AssetRow(
                asset_id="finalized:2",
                task_asset_id=2,
                kind="finalized",
                file_name=second.name,
                file_size=second.stat().st_size,
                sku_code="CGK000002",
                sku_name="商品二/3个装",
                local_path=str(second),
                status="ready",
                virtual_path=f"CGK000002/{second.name}",
            ),
        ]
    )
    hits, total = catalog.search("CGK000002", limit=20)
    assert total == 1
    assert hits[0].asset_id == "finalized:2"


def test_pack_preserves_product_names_groups_multi_image_and_keeps_duplicates(settings):
    root = settings.library_root
    product_name = "HSC36004——蔡谦-常规水晶标-喜字酒杯款-直径45cm"
    product = root / "水晶标" / product_name
    product.mkdir(parents=True)
    (product / "第一张【25x35cm】.jpg").write_bytes(b"one")
    (product / "第二张【25x35cm】.jpg").write_bytes(b"two")
    single_name = "HSC32845——鹏-常规KT板-波西和皮普-戴帽老鼠手拿气球【55x38cm】.jpg"
    (root / single_name).write_bytes(b"single")
    catalog = Catalog(settings)
    index_library(catalog, root, settings.sync.ignore_globs)

    xlsx = settings.tmp_dir / "names.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["编码"])
    sheet.append(["HSC36004"])
    sheet.append(["HSC36004"])
    sheet.append(["HSC32845"])
    workbook.save(xlsx)
    parsed = read_excel_rows(xlsx)
    unique, duplicates = deduplicate_rows(parsed)
    assert len(parsed) == 3 and len(unique) == 2 and len(duplicates) == 1

    rules = [
        {"name": "库内素材兜底", "handler": "library_fallback"},
        {"name": "保留商品名称与尺寸", "handler": "rename_sku_sequence"},
        {"name": "生成缺失清单", "handler": "missing_report"},
        {"name": "生成选择说明", "handler": "selection_report"},
        {"name": "媒体文件快速打包", "handler": "fast_zip"},
    ]
    store = JobStore(settings)
    job = store.create(filename="names.xlsx", super_dir_name="pack", meta={"rules": rules})
    job_dir = store.job_dir(job.id)
    (job_dir / "input.xlsx").write_bytes(xlsx.read_bytes())
    claimed = store.claim_next()
    assert claimed is not None
    from asset_hub.worker.main import process_job

    process_job(store, catalog, claimed.id)
    done = store.get(claimed.id)
    assert done is not None and done.status == "done"
    import zipfile

    with zipfile.ZipFile(done.archive_path) as archive:
        names = archive.namelist()
        assert f"pack/{product_name}_1/第一张【25x35cm】.jpg" in names
        assert f"pack/{product_name}_1/第二张【25x35cm】.jpg" in names
        assert f"pack/{product_name}_2/第一张【25x35cm】.jpg" in names
        assert f"pack/{product_name}_2/第二张【25x35cm】.jpg" in names
        assert f"pack/{single_name}" in names
        assert not any("HSC36004_1" in name or "HSC36004_2" in name for name in names)
        report = archive.read("pack/素材选择说明.txt").decode("utf-8")
        assert "输入编码行：3" in report
        assert "唯一编码：2" in report
        assert "与第 2 行编码重复，保留为独立商品单位" in report
    assert done.progress["input_rows"] == 3
    assert done.progress["unique_rows"] == 2
    assert done.progress["duplicate_rows"] == 1
    assert done.progress["matched"] == 3


def test_pack_keeps_single_finalized_asset_as_a_file(settings):
    source = settings.library_root / "菲瑶-常规KT板-升小学手举牌-我是小学生啦-20-50cm.jpg"
    source.write_bytes(b"single-finalized")
    catalog = Catalog(settings)
    catalog.upsert_asset(
        AssetRow(
            asset_id="finalized:10805",
            task_asset_id=10805,
            kind="finalized",
            file_name=source.name,
            file_size=source.stat().st_size,
            sku_code="CGK000672",
            sku_name="菲瑶/常规KT板/升小学手举牌/我是小学生啦",
            local_path=str(source),
            status="ready",
        )
    )

    xlsx = settings.tmp_dir / "single-finalized.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["编码"])
    sheet.append(["CGK000672"])
    workbook.save(xlsx)
    rules = [
        {"name": "当前版本优先", "handler": "prefer_current"},
        {"name": "保留商品名称与尺寸", "handler": "rename_sku_sequence"},
        {"name": "媒体文件快速打包", "handler": "fast_zip"},
    ]
    store = JobStore(settings)
    job = store.create(filename=xlsx.name, super_dir_name="pack", meta={"rules": rules})
    job_dir = store.job_dir(job.id)
    (job_dir / "input.xlsx").write_bytes(xlsx.read_bytes())
    claimed = store.claim_next()
    assert claimed is not None
    from asset_hub.worker.main import process_job

    process_job(store, catalog, claimed.id)
    done = store.get(claimed.id)
    assert done is not None and done.status == "done"
    import zipfile

    with zipfile.ZipFile(done.archive_path) as archive:
        images = [name for name in archive.namelist() if name.lower().endswith(".jpg")]
    assert images == [f"pack/{source.name}"]


def test_pack_chooses_one_best_file_for_repeated_sku_outside_product_directory(settings):
    catalog = Catalog(settings)
    parent = settings.library_root / "冯新妮皮普和波西KT板"
    parent.mkdir(parents=True)
    candidates = [
        ("HQT03449—波西和皮普—高80cm(外发).jpg", 40.0),
        ("HQT03449—波西和皮普—高80cm(多个).jpg", 30.0),
        ("HQT03449—波西和皮普—高80cm.jpg", 20.0),
        ("HQT03449—波西和皮普—高80cm._jg", 10.0),
    ]
    for name, updated_at in candidates:
        path = parent / name
        path.write_bytes(name.encode())
        rel = path.relative_to(settings.library_root).as_posix()
        catalog.upsert_asset(
            AssetRow(
                asset_id=f"lib:{rel}",
                kind="library",
                storage_key=rel,
                file_name=name,
                original_filename=name,
                file_size=path.stat().st_size,
                local_path=str(path),
                status="ready",
                updated_at=updated_at,
                virtual_path=rel,
            )
        )

    xlsx = settings.tmp_dir / "repeated-single.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["编码"])
    sheet.append(["HQT03449"])
    sheet.append(["HQT03449"])
    workbook.save(xlsx)
    rules = [
        {"name": "库内素材兜底", "handler": "library_fallback"},
        {"name": "保留商品名称与尺寸", "handler": "rename_sku_sequence"},
        {"name": "媒体文件快速打包", "handler": "fast_zip"},
    ]
    store = JobStore(settings)
    job = store.create(filename=xlsx.name, super_dir_name="pack", meta={"rules": rules})
    job_dir = store.job_dir(job.id)
    (job_dir / "input.xlsx").write_bytes(xlsx.read_bytes())
    claimed = store.claim_next()
    assert claimed is not None
    from asset_hub.worker.main import process_job

    process_job(store, catalog, claimed.id)
    done = store.get(claimed.id)
    assert done is not None and done.status == "done"
    import zipfile

    with zipfile.ZipFile(done.archive_path) as archive:
        files = [name for name in archive.namelist() if name.lower().endswith(".jpg")]
    assert files == [
        "pack/HQT03449—波西和皮普—高80cm(外发)_1.jpg",
        "pack/HQT03449—波西和皮普—高80cm(外发)_2.jpg",
    ]


def test_current_catalog_can_initialize_while_writer_holds_lock(settings):
    catalog = Catalog(settings)
    key = str(catalog.db_path.resolve())
    Catalog._schema_initialized_paths.discard(key)

    writer = sqlite3.connect(catalog.db_path, timeout=1)
    try:
        writer.execute("BEGIN IMMEDIATE")
        current = Catalog(settings)
        assert current.get_sync_state("finalized")["ready"] is False
    finally:
        writer.rollback()
        writer.close()


def test_zip_store_for_jpg(settings, tmp_path):
    src = tmp_path / "x.jpg"
    src.write_bytes(b"\xff\xd8" + b"0" * 100)
    out = settings.tmp_dir / "t.zip"
    zip_paths([(src, "pack/x.jpg")], out)
    assert out.is_file()
    import zipfile

    with zipfile.ZipFile(out) as zf:
        info = zf.getinfo("pack/x.jpg")
        assert info.compress_type == zipfile.ZIP_STORED


def test_excel_match_and_job(settings):
    sync_once()
    cat = Catalog(settings)
    xlsx = settings.tmp_dir / "pack.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU编码", "名称"])
    ws.append(["HQT10000", "样本"])
    ws.append(["MISSING999", "无"])
    wb.save(xlsx)
    rows = read_excel_rows(xlsx)
    matched, missing = match_assets_for_rows(cat, rows)
    assert matched
    assert missing

    store = JobStore(settings)
    assert store.db_path == settings.jobs_db_path
    assert store.db_path != settings.db_path
    job = store.create(filename="pack.xlsx", super_dir_name="demo")
    job_dir = store.job_dir(job.id)
    (job_dir / "input.xlsx").write_bytes(xlsx.read_bytes())
    claimed = store.claim_next()
    assert claimed and claimed.status == "running"

    from asset_hub.worker.main import process_job

    process_job(store, cat, claimed.id)
    done = store.get(claimed.id)
    assert done.status == "done"
    assert Path(done.archive_path).is_file()
    assert done.progress["preferred_rows"] >= 1
    import zipfile

    with zipfile.ZipFile(done.archive_path) as zf:
        report_names = [name for name in zf.namelist() if name.endswith("素材选择说明.txt")]
        assert report_names
        report = zf.read(report_names[0]).decode("utf-8")
        assert "当前优选" in report
        assert "MISSING999" in report


def test_job_store_accepts_and_claims_100_concurrent_jobs(settings):
    store = JobStore(settings)

    def create_one(index: int) -> str:
        return store.create(filename=f"load-{index}.xlsx").id

    with ThreadPoolExecutor(max_workers=20) as executor:
        created = list(executor.map(create_one, range(120)))
    assert len(set(created)) == 120

    def claim_all() -> list[str]:
        claimed: list[str] = []
        while True:
            job = store.claim_next()
            if not job:
                return claimed
            claimed.append(job.id)

    with ThreadPoolExecutor(max_workers=8) as executor:
        groups = list(executor.map(lambda _index: claim_all(), range(8)))
    claimed_ids = [job_id for group in groups for job_id in group]
    assert len(claimed_ids) == 120
    assert set(claimed_ids) == set(created)


def test_job_queue_is_not_blocked_by_catalog_writer(settings):
    catalog = Catalog(settings)
    store = JobStore(settings)
    writer = sqlite3.connect(catalog.db_path, timeout=1)
    try:
        writer.execute("BEGIN IMMEDIATE")
        started = time.perf_counter()
        job = store.create(filename="independent.xlsx")
        assert time.perf_counter() - started < 0.5
        assert job.status == "queued"
    finally:
        writer.rollback()
        writer.close()


def test_job_store_migrates_legacy_catalog_jobs(settings):
    catalog = Catalog(settings)
    with catalog.connect() as connection:
        connection.execute(
            """
            INSERT INTO jobs(
              id,status,created_at,filename,progress_json,meta_json
            ) VALUES (?,?,?,?,?,?)
            """,
            ("legacy-job", "done", 1.0, "legacy.xlsx", "{}", "{}"),
        )
    if settings.jobs_db_path.exists():
        settings.jobs_db_path.unlink()
    JobStore._initialized_paths.discard(str(settings.jobs_db_path.resolve()))
    migrated = JobStore(settings).get("legacy-job")
    assert migrated is not None and migrated.status == "done"


def test_api_health_search(settings):
    from asset_hub.api.main import app

    sync_once()
    client = TestClient(app)
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["ok"] is True
    r = client.get("/api/v1/search", params={"q": "HQT10000"})
    assert r.status_code == 200
    assert r.json()["results"]
    result = r.json()["results"][0]
    assert "kind" not in result
    assert "local_path" not in result
    assert "status" not in result
    status = client.get("/api/v1/status").json()
    assert status["asset_count"] > 0
    assert status["ready_for_pack"] is True

    store = JobStore(settings)
    store.create(filename="timing.xlsx")
    claimed = store.claim_next()
    assert claimed and claimed.started_at
    jobs = client.get("/api/v1/jobs").json()["jobs"]
    assert jobs[0]["started_at"] == pytest.approx(claimed.started_at)


def test_asset_single_and_batch_download(settings):
    first = settings.library_root / "HSC40063-主图.jpg"
    second = settings.library_root / "CGK001175-详情.tif"
    first.write_bytes(b"first-image")
    second.write_bytes(b"second-image")
    catalog = Catalog(settings)
    for asset_id, path in (("upload:first", first), ("upload:second", second)):
        catalog.upsert_asset(
            AssetRow(
                asset_id=asset_id,
                kind="library",
                file_name=path.name,
                file_size=path.stat().st_size,
                local_path=str(path),
                status="ready",
            )
        )

    from asset_hub.api.main import app

    client = TestClient(app)
    single = client.get("/api/v1/asset/download", params={"id": "upload:first"})
    assert single.status_code == 200
    assert single.content == b"first-image"
    assert "attachment" in single.headers["content-disposition"]

    ticket = client.post(
        "/api/v1/assets/download-ticket",
        json={"ids": ["upload:first", "upload:second"]},
    )
    assert ticket.status_code == 200
    batch = client.get(ticket.json()["download_url"])
    assert batch.status_code == 200
    assert batch.headers["content-type"] == "application/zip"
    with zipfile.ZipFile(BytesIO(batch.content)) as archive:
        assert archive.namelist() == [first.name, second.name]
        assert archive.read(first.name) == b"first-image"
        assert archive.read(second.name) == b"second-image"


def test_create_job_allows_library_fallback_before_sync_complete(settings):
    cat = Catalog(settings)
    local = settings.library_root / "HQT40001.jpg"
    local.write_bytes(b"library")
    cat.upsert_asset(
        AssetRow(
            asset_id="lib:HQT40001.jpg",
            kind="library",
            file_name=local.name,
            sku_code="HQT40001",
            local_path=str(local),
            status="ready",
        )
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU编码", "名称"])
    ws.append(["HQT40001", "样本"])
    payload = BytesIO()
    wb.save(payload)

    from asset_hub.api.main import app

    client = TestClient(app)
    response = client.post(
        "/api/v1/jobs",
        files={
            "file": (
                "library-fallback.xlsx",
                payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    assert response.json()["job_id"]
    assert response.json()["input_rows"] == 1
    assert response.json()["unique_rows"] == 1


def test_create_job_rejects_when_library_mount_is_offline(settings, monkeypatch):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["SKU编码"])
    sheet.append(["HQT05099"])
    payload = BytesIO()
    workbook.save(payload)

    from asset_hub.api import main as api_main

    monkeypatch.setattr(api_main, "library_mount_available", lambda _settings: False)
    response = TestClient(api_main.app).post(
        "/api/v1/jobs",
        files={
            "file": (
                "offline.xlsx",
                payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 503
    assert "素材盘未挂载" in response.json()["detail"]


def test_worker_fails_closed_when_library_mount_is_offline(settings, monkeypatch):
    store = JobStore(settings)
    job = store.create(filename="offline.xlsx")
    from asset_hub.worker import main as worker_main

    monkeypatch.setattr(worker_main, "library_mount_available", lambda _settings: False)
    worker_main.process_job(store, Catalog(settings), job.id)
    failed = store.get(job.id)
    assert failed is not None and failed.status == "failed"
    assert "未生成缺失清单" in failed.error


def test_library_upload_tree_and_global_filename_dedupe(settings):
    from asset_hub.api.main import app

    client = TestClient(app)
    response = client.post(
        "/api/v1/library/upload",
        data={"target_path": "产品图/主图", "relative_paths": '["manual.png"]'},
        files=[("files", ("manual.png", b"png-data", "image/png"))],
    )
    assert response.status_code == 200
    assert response.json()["added"] == 1

    tree = client.get("/api/v1/library/tree", params={"path": "产品图/主图"})
    assert tree.status_code == 200
    assert tree.json()["files"][0]["virtual_path"] == "产品图/主图/manual.png"

    duplicate = client.post(
        "/api/v1/library/upload",
        data={"target_path": "另一个目录", "relative_paths": '["manual.png"]'},
        files=[("files", ("manual.png", b"different", "image/png"))],
    )
    assert duplicate.status_code == 409
    assert duplicate.json()["detail"]["code"] == "FILENAME_DUPLICATE"
    assert len(duplicate.json()["detail"]["duplicates"]) == 1
    assert not (settings.library_root / "另一个目录" / "manual.png").exists()


def test_library_tree_search_finds_nested_sku_from_root(settings):
    catalog = Catalog(settings)
    local = settings.library_root / "水晶标" / "HSC38018——生日套装" / "第一张【25x35cm】.jpg"
    local.parent.mkdir(parents=True)
    local.write_bytes(b"nested-image")
    catalog.upsert_asset(
        AssetRow(
            asset_id="lib:水晶标/HSC38018——生日套装/第一张【25x35cm】.jpg",
            kind="library",
            storage_key="水晶标/HSC38018——生日套装/第一张【25x35cm】.jpg",
            file_name=local.name,
            original_filename=local.name,
            file_size=local.stat().st_size,
            local_path=str(local),
            status="ready",
            virtual_path="水晶标/HSC38018——生日套装/第一张【25x35cm】.jpg",
        )
    )
    from asset_hub.api.main import app

    response = TestClient(app).get(
        "/api/v1/library/tree", params={"path": "", "q": "HSC38018"}
    )
    assert response.status_code == 200
    data = response.json()
    assert data["search_mode"] is True
    assert data["directories"] == []
    assert data["total_files"] == 1
    assert data["files"][0]["virtual_path"].startswith("水晶标/HSC38018")


def test_manifest_same_filename_different_content_requires_own_ticket(settings):
    local = settings.library_root / "SAME-NAME.jpg"
    local.write_bytes(b"manual-content")
    catalog = Catalog(settings)
    catalog.upsert_asset(
        AssetRow(
            asset_id="upload:manual",
            kind="library",
            file_name=local.name,
            original_filename=local.name,
            file_size=local.stat().st_size,
            local_path=str(local),
            status="ready",
            virtual_path="手动/SAME-NAME.jpg",
        )
    )
    item = ManifestItem(
        task_asset_id=88001,
        storage_key="tasks/88/SAME-NAME.jpg",
        file_name="SAME-NAME.jpg",
        original_filename="SAME-NAME.jpg",
        file_size=999999,
        whole_hash="",
        asset_updated_at=time.time(),
        format="jpg",
        mime_type="image/jpeg",
        group_id=88,
        revision_id=89,
        revision_mode="single",
        finalized_at=time.time(),
        task_id=90,
        task_no="T-90",
        scope_kind="sku",
        sku_code="SAME88001",
        product_name="同名素材",
        revision_item_id=91,
        sort_order=0,
        item_name="主图",
    )
    catalog.apply_finalized_manifest([item], "manifest-own-object")
    alias = catalog.get_asset_by_task_asset_id(88001)
    assert alias and alias.status == "pending"
    assert alias.local_path == ""
    assert alias.dedup_of_asset_id == ""
    assert [row.task_asset_id for row in catalog.ticket_candidates(include_nonready=True)] == [88001]

    reapplied = catalog.apply_finalized_manifest([item], "manifest-own-object-next")
    assert reapplied["changed_objects"] == 1
    assert reapplied["unchanged_objects"] == 0
    assert reapplied["changed_items"] == 0
    assert reapplied["unchanged_items"] == 1


def test_external_manifest_fast_path_only_rewrites_changed_item(settings):
    catalog = Catalog(settings)
    path_a = settings.library_root / "1/KT/SKU-A.jpg"
    path_b = settings.library_root / "1/KT/SKU-B.jpg"
    path_a.parent.mkdir(parents=True, exist_ok=True)
    path_a.write_bytes(b"aaaa")
    path_b.write_bytes(b"bbbb")

    def item(asset_id: int, name: str, modified: float) -> ExternalManifestItem:
        return ExternalManifestItem(
            external_asset_id=asset_id,
            origin_path_hash=f"{asset_id:064x}",
            relative_path=f"1/KT/{name}",
            file_name=name,
            mime_type="image/jpeg",
            file_size=4,
            storage_key=f"external/{asset_id}.jpg",
            source_modified_at=modified,
            record_updated_at=modified,
            deleted=False,
        )

    first = catalog.apply_external_manifest(
        [item(1, "SKU-A.jpg", 100), item(2, "SKU-B.jpg", 100)], "external-v1"
    )
    assert first["changed"] == 2 and first["reused"] == 2
    catalog.mark_asset_status("external:1", "ready", etag="etag-a")
    catalog.mark_asset_status("external:2", "ready", etag="etag-b")

    second = catalog.apply_external_manifest(
        [item(1, "SKU-A.jpg", 200), item(2, "SKU-B.jpg", 100)], "external-v2"
    )
    assert second["changed"] == 1
    assert second["unchanged"] == 1
    changed = catalog.get_asset("external:1")
    unchanged = catalog.get_asset("external:2")
    assert changed and changed.status == "pending" and changed.etag == ""
    assert unchanged and unchanged.status == "ready" and unchanged.etag == "etag-b"


def test_new_asset_indexing_skips_delete_scans_and_tombstones_skip_indexes(settings):
    catalog = Catalog(settings)

    class RecordingConnection:
        def __init__(self, conn):
            self.conn = conn
            self.statements: list[str] = []

        def execute(self, sql, args=()):
            self.statements.append(" ".join(sql.upper().split()))
            return self.conn.execute(sql, args)

        def executemany(self, sql, args):
            self.statements.append(" ".join(sql.upper().split()))
            return self.conn.executemany(sql, args)

    with catalog.connect() as raw:
        conn = RecordingConnection(raw)
        active = AssetRow(
            asset_id="external:9001",
            kind="external",
            file_name="HQT9001.jpg",
            storage_key="external/9001.jpg",
            file_size=10,
            status="pending",
            virtual_path="1/KT/HQT9001.jpg",
        )
        Catalog._upsert_asset(conn, active, was_existing=False)
        assert not any(statement.startswith("DELETE FROM") for statement in conn.statements)
        assert any("INSERT INTO ASSETS_FTS" in statement for statement in conn.statements)

        conn.statements.clear()
        Catalog._upsert_asset(conn, active, was_existing=True)
        assert any("DELETE FROM SKU_TOKENS" in statement for statement in conn.statements)
        assert any("DELETE FROM ASSETS_FTS" in statement for statement in conn.statements)
        assert any("DELETE FROM ASSET_NAME_CLAIMS" in statement for statement in conn.statements)

        conn.statements.clear()
        tombstone = AssetRow(
            asset_id="external:9002",
            kind="external",
            file_name="old.jpg",
            status="tombstone",
            deleted=1,
            virtual_path="1/KT/old.jpg",
        )
        Catalog._upsert_asset(conn, tombstone, was_existing=False)
        assert not any(statement.startswith("DELETE FROM") for statement in conn.statements)
        assert not any("INSERT INTO ASSETS_FTS" in statement for statement in conn.statements)
        assert not any("INSERT OR IGNORE INTO SKU_TOKENS" in statement for statement in conn.statements)


def test_pack_rule_crud_and_job_snapshot(settings):
    rules = PackRuleStore(settings)
    defaults = rules.list(enabled_only=True)
    assert {rule.handler for rule in defaults} >= {
        "group_by_order",
        "repeat_quantity",
        "missing_report",
    }
    custom = rules.create(
        name="人工复核说明",
        description="结果交付前人工复核",
        handler="note",
    )
    changed = rules.update(
        custom.id,
        name="人工终检",
        description=custom.description,
        handler="note",
        enabled=True,
        sort_order=900,
        config={},
    )
    assert changed and changed.name == "人工终检"
    snapshot = rules.snapshots([custom.id])
    assert snapshot[0]["handler"] == "note"
    assert rules.delete(custom.id)


def test_eve35_columns_and_selected_pack_rules(settings):
    xlsx = settings.tmp_dir / "eve35.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["订单号", "SKU", "数量", "地址", "关键词"])
    ws.append([1001, "HQT10001", 2, "上海*", "front"])
    wb.save(xlsx)
    rows = read_excel_rows(xlsx)
    assert rows[0].order_id == "1001"
    assert rows[0].quantity == 2
    assert rows[0].address == "上海*"
    assert rows[0].keyword == "front"
